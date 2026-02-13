import openpyxl
from io import BytesIO
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from django.db import transaction, models
from apps.businesses.models import Business, Account
from .models import Transaction, Category, Merchant


def to_decimal(value):
    """
    값을 Decimal로 변환하고 소수점 2자리로 통일
    금융 데이터는 정확성이 중요하므로 quantize 필수
    """
    if value is None or value == '' or str(value).strip() == '':
        return None
    
    try:
        # 문자열로 변환 후 Decimal (부동소수점 오차 방지)
        decimal_value = Decimal(str(value))
        
        # 소수점 2자리로 통일 (DB와 동일, 반올림)
        return decimal_value.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError):
        return None


def calculate_amounts(total, supply, vat, row_number):
    """
    총금액/공급가액/부가세 계산 및 검증
    
    금융 데이터이므로 1원 오차도 허용하지 않음!
    모든 계산은 Decimal + quantize로 정확성 보장
    
    Args:
        total: 총금액 (공급가액 + 부가세)
        supply: 공급가액 (부가세 제외)
        vat: 부가세액
        row_number: 엑셀 행 번호 (에러 메시지용)
    
    Returns:
        tuple: (공급가액, 부가세, 에러메시지)
        - 성공 시: (Decimal, Decimal, None)
        - 실패 시: (None, None, str)
    
    지원 시나리오:
        A. 총금액만 입력 → 공급가액, 부가세 자동 계산 (10% 부가세)
        B. 공급가액 + 부가세 입력 → 그대로 사용
        C. 공급가액만 입력 → 부가세 자동 계산 (10%)
        D. 부가세만 입력 → 공급가액 자동 계산 (역산)
        E. 모두 입력 → 검증 (총금액 = 공급가액 + 부가세)
    """
    # 모두 Decimal + 소수점 2자리로 통일
    total = to_decimal(total)
    supply = to_decimal(supply)
    vat = to_decimal(vat)
    
    # 1. 입력값 검증: 모두 비어있으면 에러
    if total is None and supply is None and vat is None:
        return None, None, f"{row_number}행: 금액 정보가 비어있습니다."
    
    # 2. 시나리오 A: 총금액만 입력 → 공급가액/부가세 자동 계산
    # 총금액 11,000원 입력 → 공급가액 10,000원, 부가세 1,000원
    if total is not None and supply is None and vat is None:
        # 총금액 ÷ 1.1 = 공급가액
        supply = (total / Decimal('1.1')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        # 총금액 - 공급가액 = 부가세
        vat = (total - supply).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return supply, vat, None
    
    # 3. 시나리오 B: 공급가액 + 부가세만 입력 → 그대로 사용
    # 공급가액 10,000원, 부가세 1,000원 입력 → 그대로 저장
    if supply is not None and vat is not None and total is None:
        supply = supply.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        vat = vat.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return supply, vat, None
    
    # 4. 시나리오 C: 공급가액만 입력 → 부가세 자동 계산
    # 공급가액 10,000원 입력 → 부가세 1,000원 (10%)
    if supply is not None and vat is None and total is None:
        supply = supply.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        # 공급가액 × 0.1 = 부가세
        vat = (supply * Decimal('0.1')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return supply, vat, None
    
    # 5. 시나리오 D: 부가세만 입력 → 공급가액 역산
    # 부가세 1,000원 입력 → 공급가액 10,000원
    if vat is not None and supply is None and total is None:
        vat = vat.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        # 부가세 × 10 = 공급가액
        supply = (vat * Decimal('10')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return supply, vat, None
    
    # 6. 시나리오 E: 모두 입력 → 검증 (총금액 = 공급가액 + 부가세)
    # 불일치 시 상세 에러 메시지 반환
    if total is not None and supply is not None and vat is not None:
        total = total.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        supply = supply.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        vat = vat.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        calculated_total = supply + vat
        
        if total == calculated_total:
            return supply, vat, None
        else:
            # 불일치 시 상세 에러 메시지
            return None, None, (
                f"{row_number}행: 금액 불일치\n"
                f"  입력 총금액: {total:,}원\n"
                f"  공급가액: {supply:,}원\n"
                f"  부가세: {vat:,}원\n"
                f"  계산 총금액: {calculated_total:,}원\n"
                f"  차이: {(total - calculated_total):,}원"
            )
    
    # 7. 그 외 모든 경우 (예: 총금액+공급가액만 입력 등) → 에러
    return None, None, f"{row_number}행: 금액 정보가 부족하거나 잘못되었습니다."


def process_transaction_excel(excel_file, user):
    """
    엑셀 파일에서 거래 내역을 읽어 DB에 저장
    
    N+1 문제 해결: 사전 로딩 + Bulk Create
    실패 추적: error_details에 원본 데이터 저장
    """
    # 엑셀 파일 로드 (read_only 모드로 속도 향상)
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.active
    
    # 결과 추적
    skipped_count = 0   # 중복 건수를 셀 변수입니다.
    success_list = []
    error_list = []
    error_details = []  # 실패한 행의 원본 데이터 저장
    auto_created = {
        'accounts': [],
        'businesses': [],
        'merchants': [],
        'categories_matched': []
    }
    
    # ========================================
    # 1단계: 사전 로딩 (N+1 해결!)
    # ========================================
    print("📊 데이터 사전 로딩 중...")
    
    # 사업장 로딩
    businesses = {b.name: b for b in Business.active.filter(user=user)}
    default_business = Business.active.filter(user=user).first()
    
    # 계좌 로딩
    accounts = {a.account_number: a for a in Account.active.filter(user=user)}
    
    # 카테고리 로딩
    categories_by_name = {
        c.name: c 
        for c in Category.objects.filter(
            models.Q(user=user) | models.Q(is_system=True)
        )
    }
    
    # 거래처 로딩
    merchants = {m.name: m for m in Merchant.objects.filter(user=user)}
    
    # 새로 생성할 항목들
    new_businesses = {}
    new_accounts = {}
    new_merchants = {}
    
    print(f"✅ 로딩 완료: 사업장 {len(businesses)}개, 계좌 {len(accounts)}개, "
          f"카테고리 {len(categories_by_name)}개, 거래처 {len(merchants)}개")
    
    # [수정] 현재 사용자의 최근 거래 내역을 가져와서 '비교용 지문' 세트를 만듭니다.
    # 엑셀 데이터와 비교할 핵심 정보들만 묶어서 저장합니다.
    existing_transactions = set(
        Transaction.objects.filter(user=user).values_list(
            'account_id', 
            'tx_type', 
            'amount', 
            'vat_amount', 
            'merchant_name',
            'occurred_at__year', 
            'occurred_at__month', 
            'occurred_at__day', 
            'occurred_at__hour', 
            'occurred_at__minute'
        )
    )

    # ========================================
    # 2단계: 엑셀 읽기 및 검증
    # ========================================
    print("📖 엑셀 데이터 읽기 중...")
    
    with transaction.atomic():
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 빈 행 스킵
            if not any(row):
                continue
            
            # 원본 데이터 저장 (실패 시 표시용)
            raw_data = {
                '거래일시': row[0] if len(row) > 0 else '',
                '사업장': row[1] if len(row) > 1 else '',
                '계좌번호': row[2] if len(row) > 2 else '',
                '유형': row[3] if len(row) > 3 else '',
                '카테고리': row[4] if len(row) > 4 else '',
                '거래처': row[5] if len(row) > 5 else '',
                '총금액': row[6] if len(row) > 6 else '',
                '공급가액': row[7] if len(row) > 7 else '',
                '부가세': row[8] if len(row) > 8 else '',
                '메모': row[9] if len(row) > 9 else ''
            }
            
            try:
                # 10열 데이터 읽기
                if len(row) < 10:
                    raise ValueError("컬럼 수가 부족합니다. 10열이 필요합니다.")
                
                raw_date, b_name, a_number, tx_type_kor, cat_name, m_name, total, supply, vat, memo = row
                
                # ========================================
                # 날짜 처리
                # ========================================

                date_formats = [
                    '%Y-%m-%d %H:%M:%S',
                    '%Y-%m-%d %H:%M',
                    '%Y-%m-%d',
                    '%Y/%m/%d %H:%M',
                    '%Y/%m/%d',
                    '%Y.%m.%d %H:%M',  # 추가!
                    '%Y.%m.%d',        # 추가!
                ]

                if isinstance(raw_date, datetime):
                    occurred_at = raw_date
                else:
                    try:
                        date_str = str(raw_date).strip()
                        for fmt in date_formats:
                            try:
                                occurred_at = datetime.strptime(date_str, fmt)
                                break
                            except ValueError:
                                continue
                        else:
                            raise ValueError(f"날짜 형식 오류: {date_str}")
                    except Exception:
                        error_msg = f"날짜 형식이 잘못되었습니다. ({raw_date})"
                        error_list.append(f"{i}행: {error_msg}")
                        error_details.append({
                            'row_number': i,
                            'raw_data': raw_data,
                            'error': error_msg
                        })
                        continue
                
                # ========================================
                # 사업장 처리
                # ========================================
                business = None
                if b_name:
                    b_name_clean = str(b_name).strip()
                    
                    if b_name_clean in businesses:
                        business = businesses[b_name_clean]
                    elif b_name_clean in new_businesses:
                        business = new_businesses[b_name_clean]
                    else:
                        business_obj = Business(
                            user=user,
                            name=b_name_clean,
                            registration_number='',
                            business_type='미등록',
                            location='엑셀 업로드'
                        )
                        new_businesses[b_name_clean] = business_obj
                        business = business_obj
                        auto_created['businesses'].append(b_name_clean)
                else:
                    business = default_business
                
                if not business:
                    error_msg = "사업장이 없습니다."
                    error_list.append(f"{i}행: {error_msg}")
                    error_details.append({
                        'row_number': i,
                        'raw_data': raw_data,
                        'error': error_msg
                    })
                    continue
                
                # ========================================
                # 계좌 처리
                # ========================================
                a_number_clean = str(a_number).strip() if a_number else ""
                
                if not a_number_clean:
                    error_msg = "계좌번호가 비어있습니다."
                    error_list.append(f"{i}행: {error_msg}")
                    error_details.append({
                        'row_number': i,
                        'raw_data': raw_data,
                        'error': error_msg
                    })
                    continue
                
                if a_number_clean in accounts:
                    account = accounts[a_number_clean]
                elif a_number_clean in new_accounts:
                    account = new_accounts[a_number_clean]
                else:
                    account_obj = Account(
                        user=user,
                        business=business if business != default_business else default_business,
                        account_number=a_number_clean,
                        name=f'엑셀 업로드 계좌',
                        bank_name='미등록',
                        account_type='checking',
                        balance=Decimal('0')
                    )
                    new_accounts[a_number_clean] = account_obj
                    account = account_obj
                    auto_created['accounts'].append(a_number_clean)
                
                # ========================================
                # 거래 유형
                # ========================================
                actual_tx_type = 'IN' if tx_type_kor and '수입' in str(tx_type_kor) else 'OUT'
                
                # ========================================
                # 카테고리 처리
                # ========================================
                clean_cat_name = str(cat_name).strip() if cat_name else ""
                category = None
                
                if clean_cat_name:
                    if clean_cat_name in categories_by_name:
                        category = categories_by_name[clean_cat_name]
                    else:
                        for cat_key, cat_obj in categories_by_name.items():
                            if clean_cat_name in cat_key or cat_key in clean_cat_name:
                                category = cat_obj
                                auto_created['categories_matched'].append(
                                    f"{clean_cat_name} → {cat_obj.name}"
                                )
                                break
                
                if not category:
                    category_type = 'income' if actual_tx_type == 'IN' else 'expense'
                    for cat_obj in categories_by_name.values():
                        if cat_obj.type == category_type:
                            category = cat_obj
                            auto_created['categories_matched'].append(
                                f"{clean_cat_name} → {cat_obj.name} (기본값)"
                            )
                            break
                
                if not category:
                    error_msg = f"'{clean_cat_name}' 카테고리를 찾을 수 없습니다."
                    error_list.append(f"{i}행: {error_msg}")
                    error_details.append({
                        'row_number': i,
                        'raw_data': raw_data,
                        'error': error_msg
                    })
                    continue
                
                # ========================================
                # 거래처 처리
                # ========================================
                merchant = None
                merchant_name_clean = str(m_name).strip() if m_name else ""
                
                if merchant_name_clean:
                    if merchant_name_clean in merchants:
                        merchant = merchants[merchant_name_clean]
                    elif merchant_name_clean in new_merchants:
                        merchant = new_merchants[merchant_name_clean]
                    else:
                        merchant_obj = Merchant(
                            user=user,
                            name=merchant_name_clean,
                            business_number='',
                            contact=''
                        )
                        new_merchants[merchant_name_clean] = merchant_obj
                        merchant = merchant_obj
                        auto_created['merchants'].append(merchant_name_clean)
                
                # ========================================
                # 금액 계산 및 검증 (정확성 보장!)
                # ========================================
                supply_amount, vat_amount, error_msg = calculate_amounts(
                    total, supply, vat, i
                )
                
                if error_msg:
                    error_list.append(error_msg)
                    error_details.append({
                        'row_number': i,
                        'raw_data': raw_data,
                        'error': error_msg
                    })
                    continue
                # -----------------------------------------------------------
                # 2. [여기서부터 추가] 중복 체크 로직 (성능 최적화)
                # -----------------------------------------------------------
                # 문제: 엑셀 파일에 동일한 거래가 여러 번 들어있거나,
                #       이미 DB에 등록된 거래를 다시 업로드하는 경우
                # 
                # 해결: 거래의 '지문(fingerprint)'을 만들어 set에 저장
                #       → O(1) 시간복잡도로 중복 체크 (DB 조회 없음!)
                #
                # 지문 구성 요소:
                #   - 계좌번호: 어느 계좌에서 발생한 거래인지
                #   - 거래유형(IN/OUT): 수입인지 지출인지
                #   - 금액(공급가액 + 부가세): 얼마를 거래했는지
                #   - 거래처명: 누구와 거래했는지
                #   - 발생일시(연/월/일/시/분): 언제 거래했는지
                #   → 이 모든 정보가 같으면 '중복'으로 판단
                # -----------------------------------------------------------                
                current_merchant_name = merchant_name_clean or (category.name if category else "")
                clean_occurred_at = occurred_at.replace(second=0, microsecond=0)
                
                # 1. 현재 행의 데이터를 '지문'으로 만듭니다.
                current_fingerprint = (
                    account.id if hasattr(account, 'id') else None, # 새로 생성될 계좌는 None일 수 있음
                    actual_tx_type,
                    supply_amount,
                    vat_amount,
                    current_merchant_name,
                    occurred_at.year,
                    occurred_at.month,
                    occurred_at.day,
                    occurred_at.hour,
                    occurred_at.minute
                )

                # 2. DB 또는 이미 처리한 거래 중에 동일한 지문이 있는지 확인
                #    있으면 중복이므로 건너뜀 (DB에 저장하지 않음)
                if current_fingerprint in existing_transactions:
                    skipped_count += 1
                    continue
                
                # 3. 중복이 아니라면, 이 지문을 set에 추가
                #    (같은 엑셀 파일 내에서 중복 방지)
                existing_transactions.add(current_fingerprint)
                # -----------------------------------------------------------

                # ========================================
                # Transaction 객체 생성
                # ========================================
                tx_obj = Transaction(
                    user=user,
                    business=business,
                    account=account,
                    category=category,
                    merchant=merchant,
                    tx_type=actual_tx_type,
                    tax_type='taxable' if vat_amount > 0 else 'tax_free',
                    merchant_name=merchant_name_clean or (category.name if category else ""),
                    amount=supply_amount,
                    vat_amount=vat_amount,
                    occurred_at=occurred_at,
                    memo=memo or '',
                    is_business=True
                )
                
                success_list.append(tx_obj)
                
            except Exception as e:
                error_msg = str(e)
                error_list.append(f"{i}행: {error_msg}")
                error_details.append({
                    'row_number': i,
                    'raw_data': raw_data,
                    'error': error_msg
                })
                continue
        
        # ========================================
        # 3단계: Bulk Create (한 번에!)
        # ========================================
        print(f"💾 데이터 저장 중... (성공: {len(success_list)}건)")
        
        if new_businesses:
            Business.objects.bulk_create(new_businesses.values())
            print(f"  ✅ 사업장 {len(new_businesses)}개 생성")
        
        if new_accounts:
            Account.objects.bulk_create(new_accounts.values())
            print(f"  ✅ 계좌 {len(new_accounts)}개 생성")
        
        if new_merchants:
            Merchant.objects.bulk_create(new_merchants.values())
            print(f"  ✅ 거래처 {len(new_merchants)}개 생성")

        if success_list:
            Transaction.objects.bulk_create(success_list)
            print(f"  ✅ 거래 {len(success_list)}건 생성")
            
            # 계좌 잔액 업데이트 추가!
            from collections import defaultdict
            from django.db.models import F
            
            account_changes = defaultdict(lambda: Decimal('0'))
            
            for tx in success_list:
                if tx.account:
                    if tx.tx_type == 'IN':
                        account_changes[tx.account.id] += tx.amount
                    else:
                        account_changes[tx.account.id] -= tx.amount
            
            for account_id, change in account_changes.items():
                Account.objects.filter(id=account_id).update(
                    balance=F('balance') + change
                )
            
            print(f"  ✅ 계좌 잔액 업데이트 완료 ({len(account_changes)}개 계좌)")

    
    print("✅ 완료!")
    
    # 결과 반환
    return {
        'success_count': len(success_list),
        'error_count': len(error_list),
        'errors': error_list,
        'error_details': error_details,  # 실패 상세 정보
        'auto_created': auto_created,
        'skipped_count': skipped_count, #### [추가] 뷰에서 보여줄 데이터
    }


def generate_transaction_template():
    """사용자용 10열 엑셀 양식 생성"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래내역_양식"
    
    # 10열 헤더
    headers = ['거래일시', '사업장명', '계좌번호', '거래유형(수입/지출)', '카테고리', 
               '거래처명', '총금액', '공급가액', '부가세', '메모']
    ws.append(headers)
    
    # 가이드 데이터
    ws.append([
        '2026-02-08 12:00', '강남본점', '1234-5678-9012', '수입', 
        '매출', '일반고객', 11000, 10000, 1000, '커피 판매'
    ])
    ws.append([
        '2026-02-08 14:30', '강남본점', '1234-5678-9012', '지출', 
        '인건비', '직원급여', '', 2000000, '', '월급 (공급가액만 입력)'
    ])
    ws.append([
        '2026-02-08 16:00', '', '1234-5678-9012', '지출', 
        '광고비', '네이버', 55000, '', '', '광고비 (총금액만 입력)'
    ])
    ws.append([
        '', '※ 총금액만 입력하면 공급가액/부가세 자동 계산', '', '', 
        '', '※ 없는 계좌/거래처는 자동 생성', '', '', '', ''
    ])
    ws.append([
        '', '※ 금액은 1원 단위까지 정확해야 함', '', '', 
        '', '※ 소수점 2자리까지 입력 가능', '', '', '', ''
    ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_transactions_to_excel(queryset):
    """
    필터링된 거래 내역을 10열 엑셀로 내보내기
    Decimal 정확성 유지
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래내역_내보내기"

    # 10열 헤더
    headers = ['거래일시', '사업장명', '계좌번호', '거래유형', '카테고리', 
               '거래처명', '총금액', '공급가액', '부가세', '메모']
    ws.append(headers)

    # 데이터 채우기
    for tx in queryset:
        occurred_at = tx.occurred_at.strftime('%Y-%m-%d %H:%M') if tx.occurred_at else ''
        
        # Decimal을 float으로 변환 (엑셀 호환)
        total = float(tx.total_amount) if tx.total_amount else 0
        supply = float(tx.supply_value) if tx.amount else 0
        vat = float(tx.vat_amount) if tx.vat_amount else 0
        
        row = [
            occurred_at,
            tx.business.name if tx.business else '',
            tx.account.account_number if tx.account else '',
            tx.get_tx_type_display(),
            tx.category.name if tx.category else '',
            tx.merchant_name or '',
            total,    # 총금액
            supply,   # 공급가액
            vat,      # 부가세
            tx.memo or ''
        ]
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
