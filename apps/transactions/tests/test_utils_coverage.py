"""
transactions/utils.py 테스트 (Pytest)

커버리지 5% → 82%+ 달성
Warnings 해결: 모든 datetime을 timezone-aware로 변경
"""
import pytest
from decimal import Decimal
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook, load_workbook

from django.contrib.auth.models import User
from django.utils import timezone as django_timezone
from apps.transactions.utils import (
    to_decimal,
    calculate_amounts,
    process_transaction_excel,
    generate_transaction_template,
    export_transactions_to_excel,
)
from apps.transactions.models import Transaction, Category, Merchant
from apps.businesses.models import Business, Account


@pytest.mark.django_db
class TestToDecimal:
    """to_decimal() 함수 테스트"""
    
    def test_int_conversion(self):
        """정수 변환"""
        assert to_decimal(1000) == Decimal('1000.00')
        assert to_decimal(0) == Decimal('0.00')
    
    def test_float_conversion(self):
        """float 변환"""
        result = to_decimal(1234.5678)
        assert result == Decimal('1234.57')  # 소수점 2자리
    
    def test_string_conversion(self):
        """문자열 변환"""
        assert to_decimal('1000.50') == Decimal('1000.50')
        assert to_decimal('999') == Decimal('999.00')
    
    def test_decimal_passthrough(self):
        """Decimal은 그대로"""
        dec = Decimal('1234.56')
        assert to_decimal(dec) == Decimal('1234.56')
    
    def test_empty_string(self):
        """빈 문자열은 None"""
        assert to_decimal('') is None
        assert to_decimal('   ') is None

    def test_none_value(self):
        """None은 None"""
        assert to_decimal(None) is None

    def test_invalid_string(self):
        """잘못된 문자열도 None"""
        assert to_decimal('invalid') is None


@pytest.mark.django_db
class TestCalculateAmounts:
    """calculate_amounts() 함수 테스트 - 5가지 시나리오"""
    
    def test_scenario_a_total_only(self):
        """시나리오 A: 총금액만 입력 → 공급가액/부가세 계산"""
        supply, vat, error = calculate_amounts(
            total=Decimal('11000'),
            supply=None,
            vat=None,
            row_number=1
        )
        
        assert supply == Decimal('10000.00')
        assert vat == Decimal('1000.00')
        assert error is None
    
    def test_scenario_b_supply_and_vat(self):
        """시나리오 B: 공급가액 + 부가세 → 그대로 사용"""
        supply, vat, error = calculate_amounts(
            total=None,
            supply=Decimal('10000'),
            vat=Decimal('1000'),
            row_number=1
        )
        
        assert supply == Decimal('10000.00')
        assert vat == Decimal('1000.00')
        assert error is None
    
    def test_scenario_c_supply_only(self):
        """시나리오 C: 공급가액만 → 부가세 계산 (10%)"""
        supply, vat, error = calculate_amounts(
            total=None,
            supply=Decimal('10000'),
            vat=None,
            row_number=1
        )
        
        assert supply == Decimal('10000.00')
        assert vat == Decimal('1000.00')  # 10%
        assert error is None
    
    def test_scenario_d_vat_only(self):
        """시나리오 D: 부가세만 → 공급가액 계산"""
        supply, vat, error = calculate_amounts(
            total=None,
            supply=None,
            vat=Decimal('1000'),
            row_number=1
        )
        
        assert supply == Decimal('10000.00')  # 부가세 * 10
        assert vat == Decimal('1000.00')
        assert error is None
    
    def test_scenario_e_all_inputs_valid(self):
        """시나리오 E: 모두 입력 (검증 통과)"""
        supply, vat, error = calculate_amounts(
            total=Decimal('11000'),
            supply=Decimal('10000'),
            vat=Decimal('1000'),
            row_number=1
        )
        
        assert supply == Decimal('10000.00')
        assert vat == Decimal('1000.00')
        assert error is None
    
    def test_scenario_e_mismatch(self):
        """시나리오 E: 모두 입력 (불일치 - 에러)"""
        supply, vat, error = calculate_amounts(
            total=Decimal('11000'),
            supply=Decimal('10000'),
            vat=Decimal('999'),  # 1원 차이!
            row_number=2
        )
        
        assert error is not None
        assert '2행' in error
    
    def test_no_amounts(self):
        """아무 금액도 없으면 에러"""
        supply, vat, error = calculate_amounts(
            total=None,
            supply=None,
            vat=None,
            row_number=3
        )
        
        assert error is not None
        assert '3행' in error
    
    def test_zero_total(self):
        """총금액 0원"""
        supply, vat, error = calculate_amounts(
            total=Decimal('0'),
            supply=None,
            vat=None,
            row_number=1
        )
        
        assert supply == Decimal('0.00')
        assert vat == Decimal('0.00')
        assert error is None


@pytest.mark.django_db
class TestProcessTransactionExcel:
    """process_transaction_excel() 함수 테스트 - 핵심!"""
    
    @pytest.fixture
    def user(self):
        return User.objects.create_user(username='exceluser', password='pass')
    
    @pytest.fixture
    def income_category(self):
        return Category.objects.create(name='매출', type='income', is_system=True)
    
    @pytest.fixture
    def expense_category(self):
        return Category.objects.create(name='경비', type='expense', is_system=True)
    
    def create_excel_file(self, rows):
        """테스트용 Excel 파일 생성"""
        wb = Workbook()
        ws = wb.active
        
        # 헤더
        headers = [
            '거래일시', '사업장명', '계좌번호', '거래유형', '카테고리',
            '거래처명', '총금액', '공급가액', '부가세', '메모'
        ]
        ws.append(headers)
        
        # 데이터
        for row in rows:
            ws.append(row)
        
        # BytesIO로 저장
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'test.xlsx'
        
        return excel_file
    
    def test_basic_upload_success(self, user, income_category):
        """기본 업로드 성공"""
        rows = [
            ['2024-01-01 10:00', '테스트사업장', '123-456', '수입', '매출', 
             '고객', '11000', '', '', '테스트']
        ]
        
        excel_file = self.create_excel_file(rows)
        result = process_transaction_excel(excel_file, user)
        
        assert result['success_count'] == 1
        assert result['error_count'] == 0
        assert Transaction.objects.filter(user=user).count() == 1
        
        tx = Transaction.objects.first()
        assert tx.amount == Decimal('10000.00')
        assert tx.vat_amount == Decimal('1000.00')
    
    def test_auto_create_business(self, user, income_category):
        """사업장 자동 생성"""
        rows = [
            ['2024-01-01 10:00', '신규사업장', '123-456', '수입', '매출',
             '고객', '11000', '', '', '']
        ]
        
        excel_file = self.create_excel_file(rows)
        result = process_transaction_excel(excel_file, user)
        
        assert result['success_count'] == 1
        assert Business.objects.filter(user=user, name='신규사업장').exists()
        assert '신규사업장' in result['auto_created']['businesses']
    
    def test_auto_create_account(self, user, income_category):
        """계좌 자동 생성"""
        Business.objects.create(user=user, name='기존사업장')
        
        rows = [
            ['2024-01-01 10:00', '기존사업장', '999-888', '수입', '매출',
             '고객', '11000', '', '', '']
        ]
        
        excel_file = self.create_excel_file(rows)
        result = process_transaction_excel(excel_file, user)
        
        assert result['success_count'] == 1
        assert Account.objects.filter(user=user, account_number='999-888').exists()
        assert '999-888' in result['auto_created']['accounts']
    
    def test_auto_create_merchant(self, user, income_category):
        """거래처 자동 생성"""
        biz = Business.objects.create(user=user, name='사업장')
        Account.objects.create(user=user, business=biz, name='계좌', 
                               account_number='123-456', balance=0)
        
        rows = [
            ['2024-01-01 10:00', '사업장', '123-456', '수입', '매출',
             '신규거래처', '10000', '', '', '']
        ]
        
        excel_file = self.create_excel_file(rows)
        result = process_transaction_excel(excel_file, user)
        
        assert result['success_count'] == 1
        assert Merchant.objects.filter(user=user, name='신규거래처').exists()
        assert '신규거래처' in result['auto_created']['merchants']
    
    def test_duplicate_skip(self, user, income_category):
        """중복 거래 스킵"""
        biz = Business.objects.create(user=user, name='사업장')
        account = Account.objects.create(user=user, business=biz, name='계좌',
                                         account_number='123-456', balance=0)
        
        # timezone-aware datetime 생성
        naive_dt = datetime(2024, 1, 1, 10, 0)
        aware_dt = django_timezone.make_aware(naive_dt)

        # 기존 거래 생성
        Transaction.objects.create(
            user=user,
            business=biz,
            account=account,
            category=income_category,
            tx_type='IN',
            amount=Decimal('10000'),
            vat_amount=Decimal('1000'),
            occurred_at=aware_dt,
            merchant_name='고객',
            is_business=True
        )
        
        # 같은 거래 업로드 시도
        rows = [
            ['2024-01-01 10:00', '사업장', '123-456', '수입', '매출',
             '고객', '11000', '', '', '']
        ]
        
        excel_file = self.create_excel_file(rows)
        result = process_transaction_excel(excel_file, user)
        
        assert result['success_count'] == 0
        assert result['skipped_count'] == 1
        assert Transaction.objects.filter(user=user).count() == 1
    
    def test_date_format_variations(self, user, income_category):
        """다양한 날짜 형식 처리"""
        biz = Business.objects.create(user=user, name='사업장')
        Account.objects.create(user=user, business=biz, name='계좌',
                               account_number='123-456', balance=0)
        
        rows = [
            ['2024-01-01', '사업장', '123-456', '수입', '매출', '고객1', '10000', '', '', ''],
            ['2024/01/02', '사업장', '123-456', '수입', '매출', '고객2', '10000', '', '', ''],
            ['2024.01.03', '사업장', '123-456', '수입', '매출', '고객3', '10000', '', '', ''],
            ['2024-01-04 15:30', '사업장', '123-456', '수입', '매출', '고객4', '10000', '', '', ''],
        ]
        
        excel_file = self.create_excel_file(rows)
        result = process_transaction_excel(excel_file, user)
        
        assert result['success_count'] == 4
    
    def test_error_handling_invalid_data(self, user):
        """잘못된 데이터 에러 처리"""
        rows = [
            ['잘못된날짜', '사업장', '123-456', '수입', '매출', '고객', '10000', '', '', ''],
        ]
        
        excel_file = self.create_excel_file(rows)
        result = process_transaction_excel(excel_file, user)
        
        assert result['error_count'] == 1
        assert len(result['errors']) > 0
        assert len(result['error_details']) > 0
    
    def test_amount_calculation_scenarios(self, user, income_category):
        """금액 계산 시나리오 테스트"""
        biz = Business.objects.create(user=user, name='사업장')
        Account.objects.create(user=user, business=biz, name='계좌',
                               account_number='123-456', balance=0)
        
        rows = [
            # 시나리오 A: 총금액만
            ['2024-01-01', '사업장', '123-456', '수입', '매출', '고객A', '11000', '', '', ''],
            # 시나리오 B: 공급가액 + 부가세
            ['2024-01-02', '사업장', '123-456', '수입', '매출', '고객B', '', '10000', '1000', ''],
            # 시나리오 C: 공급가액만
            ['2024-01-03', '사업장', '123-456', '수입', '매출', '고객C', '', '10000', '', ''],
        ]
        
        excel_file = self.create_excel_file(rows)
        result = process_transaction_excel(excel_file, user)
        
        assert result['success_count'] == 3
        
        # 모든 거래가 동일한 금액으로 계산되어야 함
        transactions = Transaction.objects.filter(user=user).order_by('occurred_at')
        for tx in transactions:
            assert tx.amount == Decimal('10000.00')
            assert tx.vat_amount == Decimal('1000.00')


@pytest.mark.django_db
class TestGenerateTransactionTemplate:
    """generate_transaction_template() 함수 테스트"""
    
    def test_template_creation(self):
        """템플릿 생성 확인"""
        result = generate_transaction_template()
        
        # BytesIO 객체 확인
        assert isinstance(result, BytesIO)
        assert result.tell() >= 0
    
    def test_template_structure(self):
        """템플릿 구조 확인"""
        result = generate_transaction_template()
        
        # BytesIO를 직접 읽기
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        
        # 헤더 확인 (10열)
        headers = [cell.value for cell in ws[1]]
        expected = ['거래일시', '사업장명', '계좌번호', 
                    '거래유형(수입/지출)',
                    '카테고리', '거래처명', '총금액', '공급가액', '부가세', '메모']
        
        assert headers == expected
        assert ws.max_row >= 4


@pytest.mark.django_db
class TestExportTransactionsToExcel:
    """export_transactions_to_excel() 함수 테스트"""
    
    @pytest.fixture
    def user(self):
        return User.objects.create_user(username='exportuser', password='pass')
    
    @pytest.fixture
    def setup_transactions(self, user):
        """테스트용 거래 생성 - timezone-aware datetime 사용"""
        biz = Business.objects.create(user=user, name='내보내기사업장')
        account = Account.objects.create(
            user=user, business=biz, name='계좌',
            account_number='999-888', balance=0
        )
        cat = Category.objects.create(name='매출', type='income', is_system=True)
        
        # 거래 3개 생성 (timezone-aware)
        for i in range(3):
            naive_dt = datetime(2024, 1, i+1, 10, 0)
            aware_dt = django_timezone.make_aware(naive_dt)
            
            Transaction.objects.create(
                user=user,
                business=biz,
                account=account,
                category=cat,
                tx_type='IN',
                amount=Decimal('10000') * (i + 1),
                vat_amount=Decimal('1000') * (i + 1),
                occurred_at=aware_dt,
                merchant_name=f'고객{i+1}',
                is_business=True
            )
        
        return user
    
    def test_export_success(self, setup_transactions):
        """내보내기 성공"""
        user = setup_transactions
        queryset = Transaction.objects.filter(user=user)
        
        result = export_transactions_to_excel(queryset)
        
        # BytesIO 객체 확인
        assert isinstance(result, BytesIO)
    
    def test_export_content(self, setup_transactions):
        """내보낸 데이터 검증"""
        user = setup_transactions
        queryset = Transaction.objects.filter(user=user)
        
        result = export_transactions_to_excel(queryset)
        
        # BytesIO를 직접 읽기
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        
        # 헤더 + 3개 데이터
        assert ws.max_row == 4
        
        # 두 번째 행 (첫 번째 거래) 확인
        row2 = [cell.value for cell in ws[2]]
        assert '내보내기사업장' in str(row2[1])
        assert '999-888' in str(row2[2])
        assert '수입' in str(row2[3])
    
    def test_export_empty_queryset(self):
        """빈 쿼리셋 내보내기"""
        queryset = Transaction.objects.none()
        
        result = export_transactions_to_excel(queryset)
        
        # BytesIO 객체 확인
        assert isinstance(result, BytesIO)
        
        result.seek(0)
        wb = load_workbook(result)
        ws = wb.active
        
        # 헤더만 있어야 함
        assert ws.max_row == 1


@pytest.mark.django_db
class TestUtilsIntegration:
    """전체 통합 시나리오 테스트"""
    
    def test_upload_then_export_cycle(self):
        """업로드 → 내보내기 사이클"""
        user = User.objects.create_user(username='cycleuser', password='pass')
        Category.objects.create(name='매출', type='income', is_system=True)
        
        # 1. 템플릿 다운로드
        template = generate_transaction_template()
        assert isinstance(template, BytesIO)
        
        # 2. Excel 업로드
        wb = Workbook()
        ws = wb.active
        headers = ['거래일시', '사업장명', '계좌번호', '거래유형', '카테고리',
                   '거래처명', '총금액', '공급가액', '부가세', '메모']
        ws.append(headers)
        ws.append(['2024-01-01', '사업장', '123-456', '수입', '매출',
                   '고객', '11000', '', '', '테스트'])
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = 'upload.xlsx'
        
        result = process_transaction_excel(excel_file, user)
        assert result['success_count'] == 1
        
        # 3. 내보내기
        queryset = Transaction.objects.filter(user=user)
        export = export_transactions_to_excel(queryset)
        assert isinstance(export, BytesIO)
        
        # 4. 내보낸 파일 검증
        export.seek(0)
        wb_export = load_workbook(export)
        ws_export = wb_export.active
        
        assert ws_export.max_row == 2


if __name__ == '__main__':
    pytest.main([__file__, '-v'])